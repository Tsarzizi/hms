import logging
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any, Optional

from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from psycopg2.extras import RealDictCursor

from .shared.db import get_db_cursor  # ✅ 使用优化后的连接池

logger = logging.getLogger(__name__)

bp = Blueprint(
    "doctor_workload_performance",
    __name__,
    url_prefix="/api/doctor_workload_performance",
)


# ============================================================
# 工具函数：把查询结果汇总成前端需要的数据结构
# ============================================================

def _safe_num(v) -> float:
    """安全转换为浮点数"""
    if v is None:
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def build_performance_data(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    rows 来自 m_v_workload_dep_perform_total 的查询结果，
    这里按你给的参考 SQL 字段做汇总，拼成前端需要的结构。
    """

    total = {
        "结算收入": 0.0,
        "科室直接成本": 0.0,
        "绩效总额": 0.0,

        "住院工作量点数": 0.0,
        "工作量单价_合计": 0.0,
        "工作量单价_计数": 0,
        "工作量系数_合计": 0.0,
        "工作量系数_计数": 0,
        "住院工作量绩效非手术介入": 0.0,
        "基础手术绩效": 0.0,
        "介入绩效": 0.0,
        "造影绩效": 0.0,

        "结算费用": 0.0,
        "DRG病种成本": 0.0,
        "DRG结余": 0.0,
        "DRG绩效": 0.0,
        "drg系数_合计": 0.0,
        "drg系数_计数": 0,

        "门诊工作量点数": 0.0,
        "门诊工作量绩效非手术介入": 0.0,
        "门诊基础手术绩效": 0.0,
        "门诊介入绩效": 0.0,
        "门诊造影绩效": 0.0,

        "挂号费奖励": 0.0,
        "诊察费奖励": 0.0,
        "三级手术奖励": 0.0,
        "四级手术奖励": 0.0,
        "单项奖励合计": 0.0,
    }

    for row in rows:
        total["结算收入"] += _safe_num(row.get("结算收入"))
        total["科室直接成本"] += _safe_num(row.get("科室直接成本"))
        total["绩效总额"] += _safe_num(row.get("绩效总额"))

        total["住院工作量点数"] += _safe_num(row.get("住院工作量点数"))

        wl_unit = row.get("工作量单价")
        if wl_unit is not None:
            total["工作量单价_合计"] += _safe_num(wl_unit)
            total["工作量单价_计数"] += 1

        wl_coef = row.get("工作量系数")
        if wl_coef is not None:
            total["工作量系数_合计"] += _safe_num(wl_coef)
            total["工作量系数_计数"] += 1

        total["住院工作量绩效非手术介入"] += _safe_num(
            row.get("住院工作量绩效非手术介入")
        )
        total["基础手术绩效"] += _safe_num(row.get("基础手术绩效"))
        total["介入绩效"] += _safe_num(row.get("介入绩效"))
        total["造影绩效"] += _safe_num(row.get("造影绩效"))

        # DRG 相关
        total["结算费用"] += _safe_num(row.get("结算费用"))
        total["DRG病种成本"] += _safe_num(row.get("DRG病种成本"))
        total["DRG结余"] += _safe_num(row.get("DRG结余"))
        total["DRG绩效"] += _safe_num(row.get("DRG绩效"))
        drg_coef = row.get("drg系数")
        if drg_coef is not None:
            total["drg系数_合计"] += _safe_num(drg_coef)
            total["drg系数_计数"] += 1

        # 门诊相关
        total["门诊工作量点数"] += _safe_num(row.get("门诊工作量点数"))
        total["门诊工作量绩效非手术介入"] += _safe_num(
            row.get("门诊工作量绩效非手术介入")
        )
        total["门诊基础手术绩效"] += _safe_num(row.get("门诊基础手术绩效"))
        total["门诊介入绩效"] += _safe_num(row.get("门诊介入绩效"))
        total["门诊造影绩效"] += _safe_num(row.get("门诊造影绩效"))

        # 奖励
        total["挂号费奖励"] += _safe_num(row.get("挂号费奖励"))
        total["诊察费奖励"] += _safe_num(row.get("诊察费奖励"))
        total["三级手术奖励"] += _safe_num(row.get("三级手术奖励"))
        total["四级手术奖励"] += _safe_num(row.get("四级手术奖励"))
        total["单项奖励合计"] += _safe_num(row.get("单项奖励合计"))

    # 平均值
    if total["工作量单价_计数"] > 0:
        avg_workload_unit_price = (
                total["工作量单价_合计"] / total["工作量单价_计数"]
        )
    else:
        avg_workload_unit_price = 0.0

    if total["工作量系数_计数"] > 0:
        avg_workload_coef = (
                total["工作量系数_合计"] / total["工作量系数_计数"]
        )
    else:
        avg_workload_coef = 1.0

    if total["drg系数_计数"] > 0:
        avg_drg_coef = total["drg系数_合计"] / total["drg系数_计数"]
    else:
        avg_drg_coef = 0.0

    # 组装前端需要的结构（字段名对上 DoctorWorkloadPerformance.tsx）
    performance_data = {
        "inpatientWorkload": {
            "nonSurgeryOrderPoints": round(total["住院工作量点数"], 4),
            "nonSurgeryExecutePoints": round(total["住院工作量点数"], 4),
            "workloadUnitPrice": round(avg_workload_unit_price, 2),
            "nonSurgeryPerformance": round(
                total["住院工作量绩效非手术介入"], 2
            ),
            "surgeryBasePerformance": round(total["基础手术绩效"], 2),
            "angiographyPerformance": round(total["造影绩效"], 2),
            "interventionPerformance": round(total["介入绩效"], 2),
        },
        "selfPayInpatientWorkload": {
            "nonSurgeryOrderPoints": 0,
            "nonSurgeryExecutePoints": 0,
            "workloadUnitPrice": 0,
            "nonSurgeryPerformance": 0,
            "surgeryBasePerformance": 0,
            "angiographyPerformance": 0,
            "interventionPerformance": 0,
        },
        "drgPerformance": {
            "diseaseSettlementFee": round(total["结算费用"], 2),
            "drgDiseaseCost": round(total["DRG病种成本"], 2),
            "drgBalance": round(total["DRG结余"], 2),
            "drgCoefficient": round(avg_drg_coef, 4),
            "drgPerformance": round(total["DRG绩效"], 2),
        },
        "outpatientWorkload": {
            "outpatientOrderPoints": round(total["门诊工作量点数"], 4),
            "outpatientExecutePoints": round(total["门诊工作量点数"], 4),
            "workloadUnitPrice": round(avg_workload_unit_price, 2),
            "outpatientWorkloadPerformance": round(
                total["门诊工作量绩效非手术介入"], 2
            ),
        },
        "reward": {
            "outpatientRegistrationFee": round(total["挂号费奖励"], 2),
            "outpatientConsultationFee": round(total["诊察费奖励"], 2),
            "threeLevelSurgery": round(total["三级手术奖励"], 2),
            "fourLevelSurgery": round(total["四级手术奖励"], 2),
            "qualityRewardAndSubsidy": round(total["单项奖励合计"], 2),
        },
        # totalPayable 在前端会再重新计算，这里可以不算或保留 0
        "totalPayable": round(total["绩效总额"], 2),
    }

    return performance_data


# ============================================================
# /summary  汇总接口（使用你给的参考 SQL）
# ============================================================

@bp.route("/summary", methods=["GET"])
def summary():
    """
    医生工作量与绩效 - 汇总接口
    使用 m_v_workload_dep_perform_total + 参考 SQL 字段。
    前端参数：
      - month: 2025-11（yyyy-MM）
      - dep_ids: 逗号分隔绩效科室ID（可选）
      - doctor_ids: 暂时忽略（视图是科室级）
    返回：
      { "performanceData": {...} }
    """
    try:
        month: Optional[str] = request.args.get("month")
        dep_ids_raw: Optional[str] = request.args.get("dep_ids")
        dep_ids = dep_ids_raw.split(",") if dep_ids_raw else None
        # doctor_ids 目前不参与过滤（视图是科室级）
        # doctor_ids_raw = request.args.get("doctor_ids")

        sql = """
            SELECT
                yyyy_mm,
                "绩效科室ID",
                "绩效科室名称",
                "绩效科室类别",
                "绩效科室类型",
                "同类序号",
                COALESCE("人数", 0) AS "人数",
                COALESCE("结算收入", 0) AS "结算收入",
                COALESCE("科室直接成本", 0) AS "科室直接成本",
                COALESCE("绩效总额", 0) AS "绩效总额",
                COALESCE("绩效总额", 0)
                    / NULLIF(COALESCE("人数", 0), 0) AS "人均绩效",
                COALESCE("住院工作量点数", 0) AS "住院工作量点数",
                COALESCE("工作量单价", 0) AS "工作量单价",
                COALESCE("工作量系数", 0) AS "工作量系数",
                COALESCE("住院工作量绩效非手术介入", 0)
                    AS "住院工作量绩效非手术介入",
                COALESCE("基础手术绩效", 0) AS "基础手术绩效",
                COALESCE("介入绩效", 0) AS "介入绩效",
                COALESCE("造影绩效", 0) AS "造影绩效",
                COALESCE("结算收入", 0) AS "结算费用",
                COALESCE("病种成本", 0) AS "DRG病种成本",
                COALESCE("DRG结余", 0) AS "DRG结余",
                COALESCE("drg系数", 0) AS "drg系数",
                COALESCE("DRG绩效", 0) AS "DRG绩效",
                COALESCE("门诊工作量点数", 0) AS "门诊工作量点数",
                COALESCE("门诊工作量绩效非手术介入", 0)
                    AS "门诊工作量绩效非手术介入",
                COALESCE("门诊基础手术绩效", 0)
                    AS "门诊基础手术绩效",
                COALESCE("门诊介入绩效", 0)
                    AS "门诊介入绩效",
                COALESCE("门诊造影绩效", 0)
                    AS "门诊造影绩效",
                COALESCE("挂号费奖励", 0) AS "挂号费奖励",
                COALESCE("诊察费奖励", 0) AS "诊察费奖励",
                COALESCE("三级手术奖励", 0) AS "三级手术奖励",
                COALESCE("四级手术奖励", 0) AS "四级手术奖励",
                COALESCE("单项奖励合计", 0) AS "单项奖励合计"
            FROM t_workload_dep_perform_total
            WHERE
                (%(month)s IS NULL OR yyyy_mm = %(month)s)
                AND (%(dep_ids)s IS NULL OR "绩效科室ID" = ANY (%(dep_ids)s))
            ORDER BY yyyy_mm, "绩效科室类别", "同类序号"
        """
        params = {"month": month, "dep_ids": dep_ids}

        # 使用新的连接池
        with get_db_cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        performance_data = build_performance_data(rows)
        return jsonify({"performanceData": performance_data})

    except Exception:
        logger.exception("Error in /summary")
        return jsonify({"error": "服务器内部错误"}), 500


# ============================================================
# /trend  趋势接口（近 12 个月）
# ============================================================

@bp.route("/trend", methods=["GET"])
def trend():
    """
    近 12 个月趋势数据：
    返回 [
      { "month": "2025-01", "performance": 12345, "outpatient": 2000, "surgery": 3000 },
      ...
    ]
    这里 outpatient / surgery 用绩效金额来表现（方便前端画"门诊绩效 / 手术绩效"折线）。
    """
    try:
        months = int(request.args.get("months", "12") or 12)
        dep_ids_raw: Optional[str] = request.args.get("dep_ids")
        dep_ids = dep_ids_raw.split(",") if dep_ids_raw else None
        # doctor_ids 目前不参与过滤
        # doctor_ids_raw = request.args.get("doctor_ids")

        sql = f"""
            SELECT
                yyyy_mm,
                SUM(COALESCE("绩效总额", 0)) AS total_performance,
                SUM(COALESCE("门诊工作量绩效非手术介入", 0)) AS total_outpatient_perf,
                SUM(
                    COALESCE("基础手术绩效", 0)
                    + COALESCE("门诊基础手术绩效", 0)
                    + COALESCE("介入绩效", 0)
                    + COALESCE("门诊介入绩效", 0)
                    + COALESCE("造影绩效", 0)
                    + COALESCE("门诊造影绩效", 0)
                ) AS total_surgery_perf
            FROM t_workload_dep_perform_total
            WHERE
                (%(dep_ids)s IS NULL OR "绩效科室ID" = ANY (%(dep_ids)s))
            GROUP BY yyyy_mm
            ORDER BY yyyy_mm DESC
            LIMIT {months}
        """
        params = {"dep_ids": dep_ids}

        # 使用新的连接池
        with get_db_cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        # 让前端自己补齐 12 个月 0 值，我们这里只返回有数据的月份即可
        data = [
            {
                "month": r["yyyy_mm"],
                "performance": float(r["total_performance"] or 0),
                "outpatient": float(r["total_outpatient_perf"] or 0),
                "surgery": float(r["total_surgery_perf"] or 0),
            }
            for r in rows
        ]

        return jsonify(data)

    except Exception as e:
        logger.exception("Error in /trend")
        return jsonify({"error": "服务器内部错误"}), 500


# ============================================================
# /options  下拉选项（科室 / 医生）
# ============================================================

@bp.route("/options", methods=["GET"])
def options():
    """
    返回前端下拉选项：
    {
      "departments": [{ value, label }],
      "doctors": [{ value, label }]
    }
    """
    try:
        # 科室来自 m_v_workload_doc_perform_total
        sql_deps = """
            SELECT DISTINCT
                "绩效科室ID" AS dep_id,
                "绩效科室名称" AS dep_name
            FROM m_v_workload_doc_perform_total
            WHERE "绩效科室ID" IS NOT NULL
              AND "绩效科室名称" IS NOT NULL
            ORDER BY dep_id
        """

        # 医生先从 v_workload_matrix_award 中拿（如果没有这个视图则返回空）
        sql_docs = """
            SELECT DISTINCT
                doctor_id,
                doctor_name
            FROM v_workload_matrix_award
            WHERE doctor_id IS NOT NULL
              AND doctor_name IS NOT NULL
            ORDER BY doctor_id
            LIMIT 500
        """

        # 查询科室列表
        with get_db_cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql_deps)
            dep_rows = cur.fetchall()

        # 查询医生列表（可能失败）
        try:
            with get_db_cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(sql_docs)
                doc_rows = cur.fetchall()
        except Exception:
            logger.warning(
                "v_workload_matrix_award not found, doctor list will be empty."
            )
            doc_rows = []

        departments = [
            {"value": r["dep_id"], "label": r["dep_name"]} for r in dep_rows
        ]
        doctors = [
            {"value": r["doctor_id"], "label": r["doctor_name"]}
            for r in doc_rows
        ]

        return jsonify({"departments": departments, "doctors": doctors})

    except Exception as e:
        logger.exception("Error in /options")
        return jsonify({"error": "服务器内部错误"}), 500


# ============================================================
# /export  导出当前筛选条件的数据为 Excel
# ============================================================

@bp.route("/export", methods=["GET"])
def export_excel():
    """
    导出当前筛选条件下的数据（和 /summary 使用同一套查询条件 & 字段），
    生成一个 .xlsx 文件，前端 Save / 导出 按钮直接下载。
    """
    try:
        month: Optional[str] = request.args.get("month")
        dep_ids_raw: Optional[str] = request.args.get("dep_ids")
        dep_ids = dep_ids_raw.split(",") if dep_ids_raw else None
        # doctor_ids_raw = request.args.get("doctor_ids")  # 当前未使用

        # 和 /summary 一样的 SELECT，这样导出的就是当前页面的指标数据
        sql = """
            SELECT
                yyyy_mm,
                "绩效科室ID",
                "绩效科室名称",
                "绩效科室类别",
                "绩效科室类型",
                "同类序号",
                COALESCE("人数", 0) AS "人数",
                COALESCE("结算收入", 0) AS "结算收入",
                COALESCE("科室直接成本", 0) AS "科室直接成本",
                COALESCE("绩效总额", 0) AS "绩效总额",
                COALESCE("绩效总额", 0)
                    / NULLIF(COALESCE("人数", 0), 0) AS "人均绩效",
                COALESCE("住院工作量点数", 0) AS "住院工作量点数",
                COALESCE("工作量单价", 0) AS "工作量单价",
                COALESCE("工作量系数", 0) AS "工作量系数",
                COALESCE("住院工作量绩效非手术介入", 0)
                    AS "住院工作量绩效非手术介入",
                COALESCE("基础手术绩效", 0) AS "基础手术绩效",
                COALESCE("介入绩效", 0) AS "介入绩效",
                COALESCE("造影绩效", 0) AS "造影绩效",
                COALESCE("结算收入", 0) AS "结算费用",
                COALESCE("病种成本", 0) AS "DRG病种成本",
                COALESCE("DRG结余", 0) AS "DRG结余",
                COALESCE("drg系数", 0) AS "drg系数",
                COALESCE("DRG绩效", 0) AS "DRG绩效",
                COALESCE("门诊工作量点数", 0) AS "门诊工作量点数",
                COALESCE("门诊工作量绩效非手术介入", 0)
                    AS "门诊工作量绩效非手术介入",
                COALESCE("门诊基础手术绩效", 0)
                    AS "门诊基础手术绩效",
                COALESCE("门诊介入绩效", 0)
                    AS "门诊介入绩效",
                COALESCE("门诊造影绩效", 0)
                    AS "门诊造影绩效",
                COALESCE("挂号费奖励", 0) AS "挂号费奖励",
                COALESCE("诊察费奖励", 0) AS "诊察费奖励",
                COALESCE("三级手术奖励", 0) AS "三级手术奖励",
                COALESCE("四级手术奖励", 0) AS "四级手术奖励",
                COALESCE("单项奖励合计", 0) AS "单项奖励合计"
            FROM m_v_workload_doc_perform_total
            WHERE
                (%(month)s IS NULL OR yyyy_mm = %(month)s)
                AND (%(dep_ids)s IS NULL OR "绩效科室ID" = ANY (%(dep_ids)s))
            ORDER BY yyyy_mm, "绩效科室类别", "同类序号"
        """
        params = {"month": month, "dep_ids": dep_ids}

        # 使用新的连接池
        with get_db_cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

        # 构造 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "医生工作量绩效"

        if rows:
            headers = list(rows[0].keys())
        else:
            headers = [
                "yyyy_mm",
                "绩效科室ID",
                "绩效科室名称",
                "绩效科室类别",
                "绩效科室类型",
                "同类序号",
                "人数",
                "结算收入",
                "科室直接成本",
                "绩效总额",
                "人均绩效",
                "住院工作量点数",
                "工作量单价",
                "工作量系数",
                "住院工作量绩效非手术介入",
                "基础手术绩效",
                "介入绩效",
                "造影绩效",
                "结算费用",
                "DRG病种成本",
                "DRG结余",
                "drg系数",
                "DRG绩效",
                "门诊工作量点数",
                "门诊工作量绩效非手术介入",
                "门诊基础手术绩效",
                "门诊介入绩效",
                "门诊造影绩效",
                "挂号费奖励",
                "诊察费奖励",
                "三级手术奖励",
                "四级手术奖励",
                "单项奖励合计",
            ]

        # 写表头
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = cell.font.copy(bold=True)

        # 写数据
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h))

        # 输出到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"医生工作量绩效_{month or datetime.now().strftime('%Y-%m')}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        logger.exception("Error in /export")
        return jsonify({"error": "服务器内部错误"}), 500